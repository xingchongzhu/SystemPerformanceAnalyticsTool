import re
import openpyxl
import os
from datetime import datetime
from openpyxl.styles import Font, Alignment

# 将内存单位转换为MB
def convert_to_megabytes(value_str):
    if not value_str:
        return 0
    
    # 移除空格和逗号
    value_str = value_str.strip().replace(',', '')
    
    # 检查是否有单位
    if value_str.endswith('G') or value_str.endswith('g'):
        # GB到MB的转换
        return round(float(value_str[:-1]) * 1024)
    elif value_str.endswith('M') or value_str.endswith('m'):
        # 已经是MB
        return round(float(value_str[:-1]))
    elif value_str.endswith('K') or value_str.endswith('k'):
        # KB到MB的转换
        return round(float(value_str[:-1]) / 1024, 2)
    else:
        # 没有单位，假设是MB
        try:
            return round(float(value_str))
        except ValueError:
            return 0

def process_slice(slice_content, time_pattern, mem_pattern, swap_pattern, cpu_pattern, 
                 baidu_mem_pattern, baidu_cpu_pattern,
                 timestamps, mem_matches, swap_matches, cpu_matches,
                 baidu_mem_matches, baidu_cpu_matches):
    """处理单个时间片内容，提取各种性能数据"""
    # 提取时间戳
    time_match = time_pattern.search(slice_content)
    if not time_match:
        return
    
    timestamps.append(time_match.group(1))
    
    # 提取内存数据
    mem_match = mem_pattern.search(slice_content)
    if mem_match:
        mem_matches.append(mem_match.groups())
    
    # 提取交换空间数据
    swap_match = swap_pattern.search(slice_content)
    if swap_match:
        swap_matches.append(swap_match.groups())
    
    # 提取CPU数据
    cpu_match = cpu_pattern.search(slice_content)
    if cpu_match:
        cpu_matches.append(cpu_match.groups())
    
    # 提取百度导航内存数据（支持 Total RSS by process 和 Total PSS by process）
    baidu_mem_value = '0K'
    # 优先查找 Total RSS by process
    if "Total RSS by process:" in slice_content:
        rss_start = slice_content.find("Total RSS by process:")
        rss_section = slice_content[rss_start:rss_start + 5000]
        baidu_mem_match = baidu_mem_pattern.search(rss_section)
        if baidu_mem_match:
            baidu_mem_value = baidu_mem_match.group(1)
    # 如果没找到，尝试查找 Total PSS by process
    elif "Total PSS by process:" in slice_content:
        pss_start = slice_content.find("Total PSS by process:")
        pss_section = slice_content[pss_start:pss_start + 5000]
        baidu_mem_match = baidu_mem_pattern.search(pss_section)
        if baidu_mem_match:
            baidu_mem_value = baidu_mem_match.group(1)

    baidu_mem_matches.append(baidu_mem_value)
    
    # 提取百度导航CPU数据
    baidu_cpu_match = baidu_cpu_pattern.search(slice_content)
    if baidu_cpu_match:
        baidu_cpu_matches.append(baidu_cpu_match.group(1))
    else:
        baidu_cpu_matches.append('0')

def parse_perf_log(log_file, output_dir):
    # 定义正则表达式模式
    time_pattern = re.compile(r'===== 抓取次数：\d+ \| 抓取时间：([\d\- :.]+) \| 脚本已运行：[\d分秒]+ =====')
    mem_pattern = re.compile(r'\s*Mem:\s+(\d+(?:\.\d+)?[GMK]?) total,\s+(\d+(?:\.\d+)?[GMK]?) used,\s+(\d+(?:\.\d+)?[GMK]?) free,\s+(\d+(?:\.\d+)?[GMK]?) buffers')
    swap_pattern = re.compile(r'\s*Swap:\s+(\d+(?:\.\d+)?[GMK]?) total,\s+(\d+(?:\.\d+)?[GMK]?) used,\s+(\d+(?:\.\d+)?[GMK]?) free,\s+(\d+(?:\.\d+)?[GMK]?) cached')
    cpu_pattern = re.compile(r'(\d+)%cpu\s+(\d+)%user\s+(\d+)%nice\s+(\d+)%sys\s+(\d+)%idle\s+(\d+)%iow\s+(\d+)%irq\s+(\d+)%sirq\s+(\d+)%host')
    # 优化百度导航进程的内存和CPU数据正则表达式
    baidu_mem_pattern = re.compile(r'\s*(\d{1,3}(?:,\d{3})*[K])\s*:\s*com\.baidu\.naviauto\s*\(pid')
    baidu_cpu_pattern = re.compile(r'\s*\d+\s+\S+\s+\S+\s+\S+\s+\S+\s+\S+\s+\S+\s+[RS]\s+(\d+(?:\.\d+)?)\s+.*com\.baidu\.navia')
    
    # 从每个时间片提取数据
    timestamps = []
    mem_matches = []
    swap_matches = []
    cpu_matches = []
    baidu_mem_matches = []
    baidu_cpu_matches = []
    
    # 逐行处理日志文件，减少内存占用
    current_slice = []
    in_slice = False
    
    print(f"开始处理文件: {log_file}")
    line_count = 0
    slice_count = 0
    
    with open(log_file, 'r', encoding='utf-8') as f:
        for line in f:
            line_count += 1
            
            # 检测时间片开始
            if "===== 抓取次数：" in line and "抓取时间：" in line:
                # 处理上一个时间片
                if in_slice and current_slice:
                    slice_content = ''.join(current_slice)
                    process_slice(slice_content, time_pattern, mem_pattern, swap_pattern, cpu_pattern, 
                                baidu_mem_pattern, baidu_cpu_pattern,
                                timestamps, mem_matches, swap_matches, cpu_matches,
                                baidu_mem_matches, baidu_cpu_matches)
                    slice_count += 1
                    if slice_count % 10 == 0:
                        print(f"已处理 {slice_count} 个时间片")
                    
                # 开始新的时间片
                current_slice = [line]
                in_slice = True
            elif in_slice:
                # 添加到当前时间片
                current_slice.append(line)
    
    # 处理最后一个时间片
    if in_slice and current_slice:
        slice_content = ''.join(current_slice)
        process_slice(slice_content, time_pattern, mem_pattern, swap_pattern, cpu_pattern, 
                    baidu_mem_pattern, baidu_cpu_pattern,
                    timestamps, mem_matches, swap_matches, cpu_matches,
                    baidu_mem_matches, baidu_cpu_matches)
        slice_count += 1
    
    print(f"文件处理完成，共处理 {line_count} 行，{slice_count} 个时间片")
    
    # 确保所有数据列表长度一致
    min_length = min(len(timestamps), len(mem_matches), len(swap_matches), len(cpu_matches),
                    len(baidu_mem_matches), len(baidu_cpu_matches))
    timestamps = timestamps[:min_length]
    mem_matches = mem_matches[:min_length]
    swap_matches = swap_matches[:min_length]
    cpu_matches = cpu_matches[:min_length]
    baidu_mem_matches = baidu_mem_matches[:min_length]
    baidu_cpu_matches = baidu_cpu_matches[:min_length]
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    
    # 创建内存数据工作表
    ws_mem = wb.active
    ws_mem.title = "内存数据"
    mem_headers = ["抓取时间", "总内存(M)", "已用内存(M)", "可用内存(M)", "缓冲区(M)", "百度导航内存(M)"]
    for col, header in enumerate(mem_headers, 1):
        cell = ws_mem.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    # 写入内存数据
    for row, (timestamp_str, mem, baidu_mem) in enumerate(zip(timestamps, mem_matches, baidu_mem_matches), 2):
        try:
            # 将时间戳字符串转换为datetime对象
            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S.%f")
        except ValueError:
            timestamp = timestamp_str  # 如果转换失败，使用原始字符串
        ws_mem.cell(row=row, column=1, value=timestamp)
        # 设置时间格式为包含秒数
        ws_mem.cell(row=row, column=1).number_format = 'yyyy-mm-dd hh:mm:ss'
        
        # 转换内存数据单位为MB
        mem_total_mb = convert_to_megabytes(mem[0])
        mem_used_mb = convert_to_megabytes(mem[1])
        mem_free_mb = convert_to_megabytes(mem[2])
        mem_buffers_mb = convert_to_megabytes(mem[3])
        baidu_mem_mb = convert_to_megabytes(baidu_mem)
        
        ws_mem.cell(row=row, column=2, value=mem_total_mb)
        ws_mem.cell(row=row, column=3, value=mem_used_mb)
        ws_mem.cell(row=row, column=4, value=mem_free_mb)
        ws_mem.cell(row=row, column=5, value=mem_buffers_mb)
        ws_mem.cell(row=row, column=6, value=baidu_mem_mb)
    
    # 创建交换空间数据工作表
    ws_swap = wb.create_sheet(title="交换空间数据")
    swap_headers = ["抓取时间", "总交换空间(M)", "已用交换空间(M)", "可用交换空间(M)", "缓存(M)"]
    for col, header in enumerate(swap_headers, 1):
        cell = ws_swap.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    # 写入交换空间数据
    for row, (timestamp_str, swap) in enumerate(zip(timestamps, swap_matches), 2):
        try:
            # 将时间戳字符串转换为datetime对象
            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S.%f")
        except ValueError:
            timestamp = timestamp_str  # 如果转换失败，使用原始字符串
        ws_swap.cell(row=row, column=1, value=timestamp)
        # 设置时间格式为包含秒数
        ws_swap.cell(row=row, column=1).number_format = 'yyyy-mm-dd hh:mm:ss'
        
        # 转换交换空间数据单位为MB
        swap_total_mb = convert_to_megabytes(swap[0])
        swap_used_mb = convert_to_megabytes(swap[1])
        swap_free_mb = convert_to_megabytes(swap[2])
        swap_cached_mb = convert_to_megabytes(swap[3])
        
        ws_swap.cell(row=row, column=2, value=swap_total_mb)
        ws_swap.cell(row=row, column=3, value=swap_used_mb)
        ws_swap.cell(row=row, column=4, value=swap_free_mb)
        ws_swap.cell(row=row, column=5, value=swap_cached_mb)
    
    # 创建CPU数据工作表
    ws_cpu = wb.create_sheet(title="CPU数据")
    cpu_headers = ["抓取时间", "CPU总使用率(%)", "用户态CPU(%)", "Nice CPU(%)", "系统态CPU(%)", "空闲CPU(%)", "IO等待(%)", "硬中断(%)", "软中断(%)", "Host CPU(%)", "百度导航CPU(%)"]
    for col, header in enumerate(cpu_headers, 1):
        cell = ws_cpu.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    # 写入CPU数据
    for row, (timestamp_str, cpu, baidu_cpu) in enumerate(zip(timestamps, cpu_matches, baidu_cpu_matches), 2):
        try:
            # 将时间戳字符串转换为datetime对象
            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S.%f")
        except ValueError:
            timestamp = timestamp_str  # 如果转换失败，使用原始字符串
        ws_cpu.cell(row=row, column=1, value=timestamp)
        # 设置时间格式为包含秒数
        ws_cpu.cell(row=row, column=1).number_format = 'yyyy-mm-dd hh:mm:ss'
        
        for col, value in enumerate(cpu, 2):
            try:
                # 将CPU数值转换为整数
                numeric_value = int(value)
                ws_cpu.cell(row=row, column=col, value=numeric_value)
            except ValueError:
                # 如果转换失败，使用原始值
                ws_cpu.cell(row=row, column=col, value=value)
        
        # 写入百度导航CPU数据
        try:
            # baidu_cpu是一个字符串，直接转换为浮点数
            baidu_cpu_value = float(baidu_cpu)
            ws_cpu.cell(row=row, column=11, value=baidu_cpu_value)
        except (ValueError, IndexError):
            # 如果转换失败或索引错误，写入0
            ws_cpu.cell(row=row, column=11, value=0)
    
    # 调整所有工作表的列宽
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
    
    # 保存Excel文件
    log_filename = os.path.basename(log_file)
    excel_file = os.path.join(output_dir, log_filename.replace('.log', '.xlsx'))
    wb.save(excel_file)
    
    # 自动生成交互式HTML报告
    try:
        import subprocess
        import sys
        # 调用交互式报告生成脚本，并传递当前Excel文件路径
        script_path = os.path.join(os.path.dirname(__file__), 'generate_interactive_report.py')
        subprocess.run([sys.executable, script_path, '--excel', excel_file], check=True)
        print(f"已为 {log_filename} 生成交互式HTML报告")
    except Exception as e:
        print(f"生成HTML报告时出错: {str(e)}")
    
    return excel_file

if __name__ == "__main__":
    # 设置日志目录和输出目录
    log_dir = "./日志"
    output_dir = "./output"
    
    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)
    
    # 遍历日志目录下所有.log文件
    log_files = [f for f in os.listdir(log_dir) if f.endswith('.log')]
    
    for log_file in log_files:
        log_file_path = os.path.join(log_dir, log_file)
        # 检查是否已解析
        excel_file_path = os.path.join(output_dir, log_file.replace('.log', '.xlsx'))
        if os.path.exists(excel_file_path):
            continue  # 已解析，跳过
        
        # 解析日志文件
        try:
            parse_perf_log(log_file_path, output_dir)
        except Exception as e:
            print(f"解析文件 {log_file} 时出错: {str(e)}")
    
    print("所有未解析的日志文件已处理完成！")