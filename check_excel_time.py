import openpyxl
import os
from datetime import datetime

# 查找最新生成的Excel文件
output_dir = "./output"
files = [f for f in os.listdir(output_dir) if f.endswith('.xlsx') and not f.startswith('~$')]
if not files:
    print("没有找到Excel文件")
    exit()

# 获取最新的Excel文件
excel_file = sorted(files, key=lambda x: os.path.getmtime(os.path.join(output_dir, x)))[-1]
excel_path = os.path.join(output_dir, excel_file)

# 打开Excel文件
wb = openpyxl.load_workbook(excel_path)
ws = wb["内存数据"]

# 获取行数
max_row = ws.max_row
print(f"Excel文件有 {max_row-1} 行数据")

# 打印前几个时间戳
print("\n前5个时间戳:")
for row in range(2, min(7, max_row+1)):
    cell = ws.cell(row=row, column=1)
    value = cell.value
    print(f"  行{row}: {value} (类型: {type(value).__name__})")

# 检查时间戳是否递增
print("\n检查时间戳是否递增:")
prev_time = None
all_same = True
for row in range(2, min(20, max_row+1)):
    cell = ws.cell(row=row, column=1)
    curr_time = cell.value
    if prev_time is None:
        prev_time = curr_time
    else:
        if curr_time != prev_time:
            all_same = False
            print(f"  发现不同时间戳: 行{row-1} = {prev_time}, 行{row} = {curr_time}")
            break
        prev_time = curr_time

if all_same:
    print("  所有检查的时间戳都相同！")
else:
    print("  时间戳是递增的")

# 检查实际的数值表示
print("\n时间戳的数值表示:")
for row in range(2, min(7, max_row+1)):
    cell = ws.cell(row=row, column=1)
    value = cell.value
    if isinstance(value, datetime):
        # Excel存储datetime为自1900-01-01以来的天数
        # 转换为Excel的数值表示
        delta = value - datetime(1899, 12, 30)  # Excel的基准日期
        excel_value = delta.days + delta.seconds / (24 * 3600)
        print(f"  行{row}: {value} = Excel数值 {excel_value:.6f}")
    else:
        print(f"  行{row}: {value} (不是datetime类型)")

wb.close()
